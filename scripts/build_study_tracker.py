#!/usr/bin/env python3
"""Generate a Study Progress Tracker Excel workbook.

Sheets:
  - Dashboard : live summary (overall progress, status counts, hours) driven by formulas
  - Tracker   : main topic-by-topic table with dropdowns + conditional formatting
  - Schedule  : simple weekly study planner
  - Lists     : (hidden) source values for the dropdowns

Run:  python3 scripts/build_study_tracker.py
Out:  Study_Progress_Tracker.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter

OUT = "Study_Progress_Tracker.xlsx"

# ---- palette -------------------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E1F2"
LIGHTER = "EAF0FB"
GREEN = "C6EFCE"
GREEN_TXT = "006100"
YELLOW = "FFEB9C"
YELLOW_TXT = "9C6500"
RED = "FFC7CE"
RED_TXT = "9C0006"
GREY = "808080"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUSES = ["Not Started", "In Progress", "Completed", "On Hold", "Revising"]
PRIORITIES = ["High", "Medium", "Low"]
CONFIDENCE = ["1 - Shaky", "2 - Weak", "3 - OK", "4 - Good", "5 - Strong"]

wb = Workbook()

# =========================================================================
# Lists sheet (source for validations)
# =========================================================================
ws_lists = wb.active
ws_lists.title = "Lists"
ws_lists["A1"] = "Status"
ws_lists["B1"] = "Priority"
ws_lists["C1"] = "Confidence"
for i, v in enumerate(STATUSES, start=2):
    ws_lists.cell(row=i, column=1, value=v)
for i, v in enumerate(PRIORITIES, start=2):
    ws_lists.cell(row=i, column=2, value=v)
for i, v in enumerate(CONFIDENCE, start=2):
    ws_lists.cell(row=i, column=3, value=v)
ws_lists.sheet_state = "hidden"

# =========================================================================
# Tracker sheet
# =========================================================================
ws = wb.create_sheet("Tracker")
HEADERS = [
    "Subject", "Topic / Chapter", "Priority", "Status",
    "Planned Date", "Deadline", "Hours Planned", "Hours Done",
    "Progress %", "Confidence", "Notes",
]
FIRST_ROW = 4          # first data row
LAST_ROW = 203         # rows reserved for data (200 rows)

# Title band
ws.merge_cells("A1:K1")
t = ws["A1"]
t.value = "STUDY PROGRESS TRACKER"
t.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
t.fill = PatternFill("solid", fgColor=NAVY)
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:K2")
s = ws["A2"]
s.value = "Fill in your subjects and topics below. Status, Priority and Confidence are dropdowns. Progress % auto-fills from Status (override any cell to a manual number if you like)."
s.font = Font(name="Calibri", size=9, italic=True, color="595959")
s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 22

# Header row
for c, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
ws.row_dimensions[3].height = 28

# Sample rows so the template is immediately usable
samples = [
    ["Mathematics", "Algebra - Quadratic Equations", "High", "Completed", "2026-06-01", "2026-06-10", 6, 6, None, "4 - Good", "Felt solid on practice set."],
    ["Mathematics", "Calculus - Integration", "High", "In Progress", "2026-06-08", "2026-06-20", 8, 3, None, "3 - OK", "Revisit by-parts."],
    ["Physics", "Thermodynamics", "Medium", "Not Started", "2026-06-15", "2026-06-25", 5, 0, None, "1 - Shaky", ""],
    ["Chemistry", "Organic - Reaction Mechanisms", "High", "Revising", "2026-06-05", "2026-06-18", 7, 5, None, "3 - OK", "Make flashcards."],
    ["Biology", "Genetics", "Low", "On Hold", "2026-06-20", "2026-07-01", 4, 1, None, "2 - Weak", "After exams week."],
]

# Progress formula: auto from Status unless user types a number.
def progress_formula(r):
    d = f"D{r}"  # Status
    return (
        f'=IF($A{r}="","",'
        f'IF({d}="Completed",1,'
        f'IF({d}="Revising",0.9,'
        f'IF({d}="In Progress",IF($G{r}>0,MIN(1,$H{r}/$G{r}),0.5),'
        f'IF({d}="On Hold",0.25,0)))))'
    )

for idx in range(FIRST_ROW, LAST_ROW + 1):
    sample = samples[idx - FIRST_ROW] if (idx - FIRST_ROW) < len(samples) else None
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=idx, column=c)
        cell.border = BORDER
        cell.alignment = Alignment(
            horizontal="center" if c in (3, 4, 5, 6, 7, 8, 9, 10) else "left",
            vertical="center",
        )
        if sample and c <= len(sample) and c != 9:
            cell.value = sample[c - 1]
        if c in (5, 6):  # dates
            cell.number_format = "yyyy-mm-dd"
        if c in (7, 8):  # hours
            cell.number_format = "0.0"
        if c == 9:        # progress %
            cell.value = progress_formula(idx)
            cell.number_format = "0%"
    # zebra striping
    if (idx - FIRST_ROW) % 2 == 1:
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=idx, column=c).fill = PatternFill("solid", fgColor=LIGHTER)

# Column widths
widths = {"A": 16, "B": 34, "C": 10, "D": 13, "E": 13, "F": 13,
          "G": 13, "H": 12, "I": 11, "J": 13, "K": 34}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A4"
ws.sheet_view.showGridLines = False

# --- Data validations ----------------------------------------------------
dv_status = DataValidation(type="list", formula1="=Lists!$A$2:$A$6", allow_blank=True)
dv_priority = DataValidation(type="list", formula1="=Lists!$B$2:$B$4", allow_blank=True)
dv_conf = DataValidation(type="list", formula1="=Lists!$C$2:$C$6", allow_blank=True)
ws.add_data_validation(dv_status)
ws.add_data_validation(dv_priority)
ws.add_data_validation(dv_conf)
dv_priority.add(f"C{FIRST_ROW}:C{LAST_ROW}")
dv_status.add(f"D{FIRST_ROW}:D{LAST_ROW}")
dv_conf.add(f"J{FIRST_ROW}:J{LAST_ROW}")

# --- Conditional formatting ---------------------------------------------
status_rng = f"D{FIRST_ROW}:D{LAST_ROW}"
ws.conditional_formatting.add(status_rng,
    CellIsRule(operator="equal", formula=['"Completed"'],
               fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_TXT)))
ws.conditional_formatting.add(status_rng,
    CellIsRule(operator="equal", formula=['"In Progress"'],
               fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_TXT)))
ws.conditional_formatting.add(status_rng,
    CellIsRule(operator="equal", formula=['"Not Started"'],
               fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_TXT)))
ws.conditional_formatting.add(status_rng,
    CellIsRule(operator="equal", formula=['"On Hold"'],
               fill=PatternFill("solid", fgColor="D9D9D9"), font=Font(color="404040")))
ws.conditional_formatting.add(status_rng,
    CellIsRule(operator="equal", formula=['"Revising"'],
               fill=PatternFill("solid", fgColor="BDD7EE"), font=Font(color="1F3864")))

# Priority colour
pri_rng = f"C{FIRST_ROW}:C{LAST_ROW}"
ws.conditional_formatting.add(pri_rng,
    CellIsRule(operator="equal", formula=['"High"'], font=Font(color=RED_TXT, bold=True)))

# Progress colour scale
ws.conditional_formatting.add(f"I{FIRST_ROW}:I{LAST_ROW}",
    ColorScaleRule(start_type="num", start_value=0, start_color=RED,
                   mid_type="num", mid_value=0.5, mid_color=YELLOW,
                   end_type="num", end_value=1, end_color=GREEN))

# Overdue deadline highlight (deadline passed & not completed)
ws.conditional_formatting.add(f"F{FIRST_ROW}:F{LAST_ROW}",
    FormulaRule(formula=[f'AND($F{FIRST_ROW}<>"",$F{FIRST_ROW}<TODAY(),$D{FIRST_ROW}<>"Completed")'],
                fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_TXT, bold=True)))

# =========================================================================
# Dashboard sheet
# =========================================================================
dash = wb.create_sheet("Dashboard", 0)
dash.sheet_view.showGridLines = False
for col, w in {"A": 3, "B": 26, "C": 16, "D": 4, "E": 26, "F": 16, "G": 3}.items():
    dash.column_dimensions[col].width = w

dash.merge_cells("B2:F2")
h = dash["B2"]
h.value = "STUDY PROGRESS — DASHBOARD"
h.font = Font(size=18, bold=True, color=WHITE)
h.alignment = Alignment(horizontal="center", vertical="center")
h.fill = PatternFill("solid", fgColor=NAVY)
dash.row_dimensions[2].height = 32

dash.merge_cells("B3:F3")
sub = dash["B3"]
sub.value = "Auto-updates from the Tracker sheet"
sub.font = Font(size=10, italic=True, color="595959")
sub.alignment = Alignment(horizontal="center")

DR = f"Tracker!$D${FIRST_ROW}:$D${LAST_ROW}"   # status range
AR = f"Tracker!$A${FIRST_ROW}:$A${LAST_ROW}"   # subject range
IR = f"Tracker!$I${FIRST_ROW}:$I${LAST_ROW}"   # progress range
GR = f"Tracker!$G${FIRST_ROW}:$G${LAST_ROW}"   # hours planned
HR = f"Tracker!$H${FIRST_ROW}:$H${LAST_ROW}"   # hours done

def kpi(cell_label, cell_val, label, formula, numfmt="0", fill=BLUE):
    lc = dash[cell_label]
    lc.value = label
    lc.font = Font(bold=True, color=WHITE, size=11)
    lc.fill = PatternFill("solid", fgColor=fill)
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lc.border = BORDER
    vc = dash[cell_val]
    vc.value = formula
    vc.font = Font(bold=True, size=14, color=NAVY)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = numfmt
    vc.fill = PatternFill("solid", fgColor=LIGHT)
    vc.border = BORDER

# KPI cards (left column)
kpi("B5", "C5", "Total Topics", f'=COUNTA({AR})')
kpi("B6", "C6", "Completed", f'=COUNTIF({DR},"Completed")', fill=GREEN_TXT)
kpi("B7", "C7", "In Progress", f'=COUNTIF({DR},"In Progress")', fill=YELLOW_TXT)
kpi("B8", "C8", "Not Started", f'=COUNTIF({DR},"Not Started")', fill=RED_TXT)
kpi("B9", "C9", "On Hold / Revising",
    f'=COUNTIF({DR},"On Hold")+COUNTIF({DR},"Revising")', fill=GREY)

# KPI cards (right column)
kpi("E5", "F5", "Overall Progress", f'=IFERROR(AVERAGE({IR}),0)', numfmt="0%")
kpi("E6", "F6", "Hours Planned", f'=SUM({GR})', numfmt="0.0")
kpi("E7", "F7", "Hours Done", f'=SUM({HR})', numfmt="0.0")
kpi("E8", "F8", "Hours Remaining", f'=MAX(0,SUM({GR})-SUM({HR}))', numfmt="0.0")
kpi("E9", "F9", "Overdue Topics",
    f'=SUMPRODUCT(--(Tracker!$F${FIRST_ROW}:$F${LAST_ROW}<>""),'
    f'--(Tracker!$F${FIRST_ROW}:$F${LAST_ROW}<TODAY()),'
    f'--(Tracker!$D${FIRST_ROW}:$D${LAST_ROW}<>"Completed"))', fill=RED_TXT)

# Per-subject breakdown table
dash.merge_cells("B11:F11")
bt = dash["B11"]
bt.value = "PROGRESS BY SUBJECT"
bt.font = Font(bold=True, color=WHITE, size=12)
bt.fill = PatternFill("solid", fgColor=BLUE)
bt.alignment = Alignment(horizontal="left", vertical="center", indent=1)
dash.row_dimensions[11].height = 22

sub_headers = ["Subject", "Topics", "Done", "Avg Progress"]
# place across B,C,D + merged E:F for progress
dash["B12"] = sub_headers[0]
dash["C12"] = sub_headers[1]
dash["D12"] = sub_headers[2]
dash.merge_cells("E12:F12")
dash["E12"] = sub_headers[3]
for ref in ("B12", "C12", "D12", "E12"):
    cc = dash[ref]
    cc.font = Font(bold=True, color=NAVY)
    cc.fill = PatternFill("solid", fgColor=LIGHT)
    cc.alignment = Alignment(horizontal="center", vertical="center")
    cc.border = BORDER
dash["F12"].border = BORDER

# Pull subject names; sample subjects pre-listed, rest blank for the user.
subject_seed = ["Mathematics", "Physics", "Chemistry", "Biology", "", "", "", ""]
start = 13
for i, subj in enumerate(subject_seed):
    r = start + i
    b = dash.cell(row=r, column=2, value=subj)
    b.border = BORDER
    b.alignment = Alignment(horizontal="left", indent=1)
    # Topics count
    c = dash.cell(row=r, column=3)
    c.value = f'=IF($B{r}="","",COUNTIF({AR},$B{r}))'
    c.border = BORDER
    c.alignment = Alignment(horizontal="center")
    # Done count
    d = dash.cell(row=r, column=4)
    d.value = (f'=IF($B{r}="","",COUNTIFS({AR},$B{r},{DR},"Completed"))')
    d.border = BORDER
    d.alignment = Alignment(horizontal="center")
    # Avg progress (merged E:F)
    dash.merge_cells(f"E{r}:F{r}")
    e = dash.cell(row=r, column=5)
    e.value = (f'=IF($B{r}="","",IFERROR(AVERAGEIF({AR},$B{r},{IR}),0))')
    e.number_format = "0%"
    e.border = BORDER
    e.alignment = Alignment(horizontal="center")
    dash.cell(row=r, column=6).border = BORDER

# progress colour scale on subject table
dash.conditional_formatting.add(f"E{start}:E{start+len(subject_seed)-1}",
    ColorScaleRule(start_type="num", start_value=0, start_color=RED,
                   mid_type="num", mid_value=0.5, mid_color=YELLOW,
                   end_type="num", end_value=1, end_color=GREEN))

dash.merge_cells(f"B{start+len(subject_seed)+1}:F{start+len(subject_seed)+1}")
tip = dash.cell(row=start+len(subject_seed)+1, column=2)
tip.value = "Tip: add your own subjects in column B above — the counts and progress fill in automatically."
tip.font = Font(italic=True, size=9, color="595959")

# =========================================================================
# Schedule sheet (weekly planner)
# =========================================================================
sch = wb.create_sheet("Weekly Schedule")
sch.sheet_view.showGridLines = False
days = ["Time", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
sch.merge_cells("A1:H1")
st = sch["A1"]
st.value = "WEEKLY STUDY SCHEDULE"
st.font = Font(size=16, bold=True, color=WHITE)
st.alignment = Alignment(horizontal="center", vertical="center")
st.fill = PatternFill("solid", fgColor=NAVY)
sch.row_dimensions[1].height = 28

for c, day in enumerate(days, start=1):
    cell = sch.cell(row=2, column=c, value=day)
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER

slots = ["06:00 - 08:00", "08:00 - 10:00", "10:00 - 12:00", "12:00 - 14:00",
         "14:00 - 16:00", "16:00 - 18:00", "18:00 - 20:00", "20:00 - 22:00"]
for r, slot in enumerate(slots, start=3):
    tc = sch.cell(row=r, column=1, value=slot)
    tc.font = Font(bold=True, color=NAVY)
    tc.fill = PatternFill("solid", fgColor=LIGHT)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = BORDER
    for c in range(2, 9):
        cell = sch.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if r % 2 == 1:
            cell.fill = PatternFill("solid", fgColor=LIGHTER)
    sch.row_dimensions[r].height = 26

sch.column_dimensions["A"].width = 16
for c in range(2, 9):
    sch.column_dimensions[get_column_letter(c)].width = 18
sch.freeze_panes = "B3"

# order sheets: Dashboard, Tracker, Weekly Schedule, Lists(hidden)
order = ["Dashboard", "Tracker", "Weekly Schedule", "Lists"]
wb._sheets.sort(key=lambda s: order.index(s.title))

wb.save(OUT)
print(f"Wrote {OUT}")
